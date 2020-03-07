import openpyxl
import os
import datetime

path = os.path.abspath(os.curdir)+'\\Table'
files_table = os.listdir(path)

for file in files_table:
    
    wb = openpyxl.load_workbook(filename = path+'//'+file) # открываем файл таблицы
    sheet=wb[file[0:len(file)-5]]
    
    row=2
    colomns_max =sheet.max_column
    param=[]
    while (row!=sheet.max_row+1):
        buf_string='INSERT INTO'
        buf_string+=' '+file[0:len(file)-5]+' VALUES ('
        for i in range(1,colomns_max+1):
            if isinstance(sheet.cell(row=row, column=i).value,int):
                 param.append(sheet.cell(row=row, column=i).value)
            else:
                if isinstance(sheet.cell(row=row, column=i).value,datetime.date):
                    el=sheet.cell(row=row, column=i).value
                    string_time='"'+str(el.year)
                    if (el.month>=1) &(el.month<=9):
                        string_time+='-'+'0'+str(el.month)
                    else:
                        string_time+='-'+str(el.month)
                    if (el.day>=1) &(el.day<=9):
                        string_time+='-'+'0'+str(el.day)+'"'
                        param.append(string_time)

                    else:
                        string_time+='-'+str(el.day)+'"'
                        param.append(string_time)
                else:
                    param.append('"'+sheet.cell(row=row, column=i).value+'"')
        for j in param:
            buf_string+= str(j)+','
        print(buf_string[0:len(buf_string)-1]+');')
        param=[]
        row+=1

            
       

        
    


